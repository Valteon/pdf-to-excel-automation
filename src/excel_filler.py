"""
Módulo para rellenar plantillas Excel de forma inteligente
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from typing import Dict, List, Any, Tuple
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelFiller:
    """Rellena plantillas Excel con datos extraídos de manera inteligente"""
    
    def __init__(self, template_path: str):
        """
        Inicializa el rellenador de Excel
        
        Args:
            template_path: Ruta de la plantilla Excel
        """
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None
        self.mappings = {}
        
    def load_template(self, sheet_name: str = None):
        """
        Carga la plantilla Excel
        
        Args:
            sheet_name: Nombre de la hoja (si es None usa la primera)
        """
        try:
            self.workbook = openpyxl.load_workbook(self.template_path)
            
            if sheet_name:
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
                
            logger.info(f"✓ Plantilla cargada: {self.template_path}")
            return True
        except Exception as e:
            logger.error(f"Error al cargar plantilla: {e}")
            return False
    
    def find_cells_by_pattern(self, pattern: str) -> List[Tuple[int, int]]:
        """
        Encuentra celdas que coincidan con un patrón
        
        Args:
            pattern: Patrón a buscar (ej: 'nombre', 'fecha', etc)
            
        Returns:
            Lista de tuplas (fila, columna)
        """
        matching_cells = []
        
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value and pattern.lower() in str(cell.value).lower():
                    matching_cells.append((cell.row, cell.column))
                    logger.info(f"✓ Coincidencia encontrada en {cell.coordinate}: {cell.value}")
        
        return matching_cells
    
    def map_fields(self, mapping: Dict[str, str]):
        """
        Define el mapeo entre campos de datos y celdas Excel
        
        Args:
            mapping: Diccionario con {campo_datos: celda_excel}
                    Ejemplo: {'nombre': 'B2', 'fecha': 'C3'}
        """
        self.mappings = mapping
        logger.info(f"✓ Mapeo definido: {mapping}")
    
    def auto_map_fields(self, data: Dict[str, Any]):
        """
        Intenta mapear automáticamente los campos disponibles con los de la plantilla
        
        Args:
            data: Diccionario con datos extraídos
        """
        auto_mapping = {}
        
        # Buscar coincidencias automáticas
        for key in data.keys():
            cells = self.find_cells_by_pattern(key)
            if cells:
                auto_mapping[key] = cells[0]  # Usar la primera coincidencia
        
        self.mappings = auto_mapping
        logger.info(f"✓ Mapeo automático realizado: {auto_mapping}")
        return auto_mapping
    
    def fill_cell(self, cell_reference: str, value: Any, style: bool = True):
        """
        Rellena una celda individual
        
        Args:
            cell_reference: Referencia de celda (ej: 'B2')
            value: Valor a insertar
            style: Si se deben aplicar estilos
        """
        try:
            cell = self.worksheet[cell_reference]
            cell.value = value
            
            if style:
                # Aplicar estilos básicos
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            logger.info(f"✓ Celda {cell_reference} rellenada: {value}")
        except Exception as e:
            logger.error(f"Error al rellenar celda {cell_reference}: {e}")
    
    def fill_with_data(self, data: Dict[str, Any], create_mapping: bool = True) -> bool:
        """
        Rellena la plantilla con los datos proporcionados
        
        Args:
            data: Diccionario con datos a rellenar
            create_mapping: Si True, intenta crear mapeo automático
            
        Returns:
            True si se rellenó correctamente
        """
        if not self.worksheet:
            logger.error("Primero debes cargar la plantilla")
            return False
        
        if create_mapping and not self.mappings:
            self.auto_map_fields(data)
        
        try:
            for field, cell_ref in self.mappings.items():
                if field in data:
                    value = data[field]
                    
                    # Convertir referencias de tupla a string si es necesario
                    if isinstance(cell_ref, tuple):
                        cell = self.worksheet.cell(row=cell_ref[0], column=cell_ref[1])
                    else:
                        cell = self.worksheet[cell_ref]
                    
                    cell.value = value
                    logger.info(f"✓ {field} → {cell.coordinate}: {value}")
            
            logger.info("✓ Plantilla rellenada correctamente")
            return True
        except Exception as e:
            logger.error(f"Error al rellenar plantilla: {e}")
            return False
    
    def fill_table_from_dataframe(self, df: pd.DataFrame, start_cell: str = 'A1', 
                                   include_header: bool = True):
        """
        Rellena una tabla desde un DataFrame
        
        Args:
            df: DataFrame con los datos
            start_cell: Celda inicial
            include_header: Si se incluyen los encabezados
        """
        try:
            start_row = openpyxl.utils.cell.coordinate_to_tuple(start_cell)[0]
            start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)[1]
            
            if include_header:
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = self.worksheet.cell(row=start_row, column=start_col + col_idx - 1)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", 
                                           fill_type="solid")
                start_row += 1
            
            for row_idx, row in enumerate(df.values, start_row):
                for col_idx, value in enumerate(row, start_col):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
            
            logger.info(f"✓ Tabla de {len(df)} filas insertada desde {start_cell}")
            return True
        except Exception as e:
            logger.error(f"Error al rellenar tabla: {e}")
            return False
    
    def add_formula(self, cell_reference: str, formula: str):
        """
        Añade una fórmula a una celda
        
        Args:
            cell_reference: Referencia de celda
            formula: Fórmula a insertar
        """
        try:
            cell = self.worksheet[cell_reference]
            cell.value = formula
            logger.info(f"✓ Fórmula añadida en {cell_reference}: {formula}")
        except Exception as e:
            logger.error(f"Error al añadir fórmula: {e}")
    
    def save_file(self, output_path: str = None) -> bool:
        """
        Guarda el archivo Excel rellenado
        
        Args:
            output_path: Ruta de salida (si es None usa template_path con "_filled")
            
        Returns:
            True si se guardó correctamente
        """
        if not self.workbook:
            logger.error("No hay workbook cargado")
            return False
        
        if output_path is None:
            base, ext = self.template_path.rsplit('.', 1)
            output_path = f"{base}_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
        
        try:
            self.workbook.save(output_path)
            logger.info(f"✓ Archivo guardado: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error al guardar archivo: {e}")
            return False
    
    def close(self):
        """Cierra el workbook"""
        if self.workbook:
            self.workbook.close()
            logger.info("✓ Workbook cerrado")
